import json
import math
import re
from collections import Counter
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "原始文件"

SQL_KEYWORDS = [
    "WITH",
    "SELECT",
    "FROM",
    "LEFT JOIN",
    "RIGHT JOIN",
    "INNER JOIN",
    "FULL JOIN",
    "JOIN",
    "WHERE",
    "GROUP BY",
    "HAVING",
    "UNION ALL",
    "UNION",
    "ORDER BY",
    "EXISTS",
    "NOT EXISTS",
    "ON",
    "AND",
    "OR",
]


def normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def strip_sql_comments(sql: str) -> str:
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.S)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql


def canonicalize_sql(sql: str) -> str:
    sql = strip_sql_comments(sql)
    sql = sql.replace("\u3000", " ")
    return normalize_whitespace(sql)


def split_top_level(expr: str, delimiter: str) -> List[str]:
    parts: List[str] = []
    current: List[str] = []
    depth = 0
    quote = None
    i = 0
    upper = expr.upper()
    delimiter_upper = delimiter.upper()
    while i < len(expr):
        ch = expr[i]
        if quote:
            current.append(ch)
            if ch == quote:
                quote = None
            i += 1
            continue
        if ch in ("'", '"'):
            quote = ch
            current.append(ch)
            i += 1
            continue
        if ch == "(":
            depth += 1
            current.append(ch)
            i += 1
            continue
        if ch == ")":
            depth = max(0, depth - 1)
            current.append(ch)
            i += 1
            continue
        if depth == 0 and upper.startswith(delimiter_upper, i):
            before_ok = i == 0 or not upper[i - 1].isalnum()
            after_idx = i + len(delimiter_upper)
            after_ok = after_idx >= len(expr) or not upper[after_idx].isalnum()
            if before_ok and after_ok:
                piece = "".join(current).strip()
                if piece:
                    parts.append(piece)
                current = []
                i += len(delimiter_upper)
                continue
        current.append(ch)
        i += 1
    piece = "".join(current).strip()
    if piece:
        parts.append(piece)
    return parts


def extract_clause(sql: str, start: str, ends: Sequence[str]) -> str:
    upper = sql.upper()
    start_idx = upper.find(start)
    if start_idx < 0:
        return ""
    start_idx += len(start)
    end_idx = len(sql)
    for end in ends:
        idx = upper.find(end, start_idx)
        if idx >= 0:
            end_idx = min(end_idx, idx)
    return sql[start_idx:end_idx].strip()


def dedupe_keep_order(items: Iterable[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for item in items:
        key = item.lower()
        if not item or key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def clean_identifier(name: str) -> str:
    name = re.sub(r'["`\[\]]', "", name or "")
    return name.strip().lower()


def table_base_name(name: str) -> str:
    name = clean_identifier(name)
    if "." in name:
        name = name.split(".")[-1]
    return name


def extract_tables(sql: str) -> List[str]:
    sql = canonicalize_sql(sql)
    matches = re.findall(
        r"\b(?:FROM|JOIN|UPDATE|INTO)\s+([A-Z0-9_.$#]+)",
        sql.upper(),
        flags=re.I,
    )
    return dedupe_keep_order([table_base_name(m) for m in matches])


def extract_group_by(sql: str) -> List[str]:
    clause = extract_clause(canonicalize_sql(sql), "GROUP BY", [" HAVING", " ORDER BY", " UNION", " UNION ALL"])
    if not clause:
        return []
    return [normalize_whitespace(x).lower() for x in split_top_level(clause, ",")]


def extract_order_by(sql: str) -> List[str]:
    clause = extract_clause(canonicalize_sql(sql), "ORDER BY", [" UNION", " UNION ALL"])
    if not clause:
        return []
    return [normalize_whitespace(x).lower() for x in split_top_level(clause, ",")]


def split_conditions(sql: str) -> List[str]:
    clause = extract_clause(canonicalize_sql(sql), "WHERE", [" GROUP BY", " HAVING", " ORDER BY", " UNION", " UNION ALL"])
    if not clause:
        return []
    parts = split_top_level(clause, "AND")
    normalized = []
    for part in parts:
        p = normalize_whitespace(part)
        p = re.sub(r"\b\d+\b", "?", p)
        p = re.sub(r"'[^']*'", "'?'", p)
        normalized.append(p.lower())
    return dedupe_keep_order(normalized)


def count_keyword(sql: str, pattern: str) -> int:
    return len(re.findall(pattern, sql, flags=re.I))


def extract_select_targets(sql: str) -> List[str]:
    clause = extract_clause(canonicalize_sql(sql), "SELECT", [" FROM"])
    if not clause:
        return []
    clause = re.sub(r"^\s*DISTINCT\s+", "", clause, flags=re.I)
    return [normalize_whitespace(x).lower() for x in split_top_level(clause, ",")]


def normalize_token_name(token: str) -> str:
    token = normalize_whitespace(token).lower()
    token = re.sub(r"\bas\s+[a-z0-9_.$#]+\b", "", token)
    token = re.sub(r"\b[a-z0-9_.$#]+\.", "", token)
    token = re.sub(r"'[^']*'", "'?'", token)
    token = re.sub(r"\b\d+\b", "?", token)
    return token.strip()


def extract_functions(sql: str) -> List[str]:
    funcs = re.findall(r"\b([A-Z_][A-Z0-9_]*)\s*\(", sql.upper())
    blacklist = {"SELECT", "FROM", "WHERE", "GROUP", "ORDER", "IN", "ON"}
    return dedupe_keep_order([f.lower() for f in funcs if f not in blacklist])


def format_sql_lines(sql: str) -> List[str]:
    text = strip_sql_comments(sql or "").replace("\r", "\n").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    for keyword in sorted(SQL_KEYWORDS, key=len, reverse=True):
        pattern = re.compile(rf"\b{re.escape(keyword)}\b", flags=re.I)
        text = pattern.sub(lambda m: "\n" + m.group(0).upper(), text)
    lines = [normalize_whitespace(line) for line in text.split("\n")]
    return [line for line in lines if line]


def find_line_numbers(lines: Sequence[str], keywords: Sequence[str]) -> List[int]:
    out: List[int] = []
    for idx, line in enumerate(lines, start=1):
        upper = line.upper()
        if any(k.upper() in upper for k in keywords):
            out.append(idx)
    return out


def format_line_refs(lines: Sequence[int]) -> str:
    if not lines:
        return ""
    if len(lines) > 6:
        return "、".join(str(x) for x in lines[:6]) + "..."
    return "、".join(str(x) for x in lines)


def jaccard(a: Sequence[str], b: Sequence[str]) -> float:
    sa = {x for x in a if x}
    sb = {x for x in b if x}
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def overlap_ratio(a: Sequence[str], b: Sequence[str]) -> float:
    sa = {x for x in a if x}
    sb = {x for x in b if x}
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / max(1, min(len(sa), len(sb)))


@dataclass
class SqlFeatures:
    raw_sql: str
    normalized_sql: str
    lines: List[str]
    tables: List[str]
    where_conditions: List[str]
    group_by: List[str]
    order_by: List[str]
    select_targets: List[str]
    functions: List[str]
    join_count: int
    subquery_count: int
    exists_count: int
    union_count: int
    has_group_by: bool
    has_with: bool
    aggregate_funcs: List[str]

    @classmethod
    def from_sql(cls, sql: str) -> "SqlFeatures":
        normalized = canonicalize_sql(sql)
        functions = extract_functions(normalized)
        aggregate = [f for f in functions if f in {"count", "sum", "avg", "min", "max", "decode", "nvl", "trunc"}]
        return cls(
            raw_sql=sql,
            normalized_sql=normalized,
            lines=format_sql_lines(sql),
            tables=extract_tables(normalized),
            where_conditions=split_conditions(normalized),
            group_by=extract_group_by(normalized),
            order_by=extract_order_by(normalized),
            select_targets=[normalize_token_name(x) for x in extract_select_targets(normalized)],
            functions=functions,
            join_count=count_keyword(normalized, r"\bJOIN\b"),
            subquery_count=max(0, count_keyword(normalized, r"\bSELECT\b") - 1),
            exists_count=count_keyword(normalized, r"\bEXISTS\b"),
            union_count=count_keyword(normalized, r"\bUNION\b"),
            has_group_by="GROUP BY" in normalized.upper(),
            has_with=normalized.upper().startswith("WITH ") or " WITH " in normalized.upper(),
            aggregate_funcs=dedupe_keep_order(aggregate),
        )


@dataclass
class SqlRecord:
    seq: int
    source_row: int
    sql: str
    meta: Dict[str, str]
    features: SqlFeatures = field(init=False)

    def __post_init__(self) -> None:
        self.features = SqlFeatures.from_sql(self.sql)


def load_0709_records() -> List[SqlRecord]:
    wb = openpyxl.load_workbook(SOURCE_DIR / "计划0709.xlsx", data_only=True)
    ws = wb["Sheet1"]
    headers = [cell.value for cell in ws[1]]
    rows: List[SqlRecord] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        data = dict(zip(headers, row))
        rows.append(
            SqlRecord(
                seq=len(rows) + 1,
                source_row=row_idx,
                sql=str(data.get("SQL语句") or ""),
                meta={
                    "任务编号": str(data.get("任务编号") or ""),
                    "应用场景": str(data.get("应用场景") or ""),
                    "代码位置": str(data.get("代码位置") or ""),
                    "修复情况": str(data.get("修复情况") or ""),
                    "开发人员": str(data.get("开发人员") or ""),
                },
            )
        )
    return rows


def load_0710_records() -> List[SqlRecord]:
    wb = openpyxl.load_workbook(SOURCE_DIR / "慢SQL表_2026_7_10.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    rows: List[SqlRecord] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        data = dict(zip(headers, row))
        rows.append(
            SqlRecord(
                seq=int(data.get("序号")),
                source_row=row_idx,
                sql=str(data.get("SQL语句") or ""),
                meta={
                    "来源": str(data.get("来源") or ""),
                    "指纹": str(data.get("指纹") or ""),
                    "服务": str(data.get("服务") or ""),
                    "平均响应时间(ms)": str(data.get("平均响应时间(ms)") or ""),
                    "执行次数": str(data.get("执行次数") or ""),
                },
            )
        )
    return rows


def load_0713_records() -> List[SqlRecord]:
    html = (SOURCE_DIR / "慢SQL表_2026_7_13 (1).xls").read_text(encoding="utf-8")
    df = pd.read_html(StringIO(html))[0]
    df.columns = ["序号", "来源", "指纹", "SQL语句", "平均响应时间(ms)", "执行次数", "服务", "异常类型"]
    df = df.iloc[1:].reset_index(drop=True)
    rows: List[SqlRecord] = []
    for idx, row in df.iterrows():
        rows.append(
            SqlRecord(
                seq=int(row["序号"]),
                source_row=idx + 2,
                sql=str(row["SQL语句"] or ""),
                meta={
                    "来源": str(row["来源"] or ""),
                    "指纹": str(row["指纹"] or ""),
                    "服务": str(row["服务"] or ""),
                    "平均响应时间(ms)": str(row["平均响应时间(ms)"] or ""),
                    "执行次数": str(row["执行次数"] or ""),
                },
            )
        )
    return rows


def score_pair(left: SqlFeatures, right: SqlFeatures) -> Dict[str, float]:
    table_j = jaccard(left.tables, right.tables)
    table_o = overlap_ratio(left.tables, right.tables)
    where_j = jaccard(left.where_conditions, right.where_conditions)
    group_j = jaccard(left.group_by, right.group_by)
    select_o = overlap_ratio(left.select_targets, right.select_targets)
    func_j = jaccard(left.aggregate_funcs, right.aggregate_funcs)
    shape_penalty = (
        abs(left.join_count - right.join_count) * 0.02
        + abs(left.union_count - right.union_count) * 0.04
        + abs(left.subquery_count - right.subquery_count) * 0.03
    )
    same_family_bonus = 0.06 if left.has_group_by == right.has_group_by else 0.0
    same_family_bonus += 0.04 if left.has_with == right.has_with else 0.0
    score = (
        table_j * 0.30
        + table_o * 0.20
        + where_j * 0.18
        + group_j * 0.07
        + select_o * 0.12
        + func_j * 0.09
        + min(left.exists_count, right.exists_count) / max(1, max(left.exists_count, right.exists_count)) * 0.04
        + same_family_bonus
        - shape_penalty
    )
    return {
        "score": max(0.0, min(1.0, score)),
        "table_j": table_j,
        "table_o": table_o,
        "where_j": where_j,
        "group_j": group_j,
        "select_o": select_o,
        "func_j": func_j,
    }


def describe_match(score: float, table_o: float) -> Tuple[str, str]:
    if score >= 0.72 and table_o >= 0.60:
        return "同一业务SQL", "高"
    if score >= 0.58 and table_o >= 0.45:
        return "可能同一业务SQL", "中"
    return "非同一业务SQL", "低"


def choose_candidates(base: SqlRecord, candidates: Sequence[SqlRecord], limit: int = 3) -> List[Tuple[SqlRecord, Dict[str, float]]]:
    scored = [(cand, score_pair(base.features, cand.features)) for cand in candidates]
    scored.sort(
        key=lambda item: (
            item[1]["score"],
            item[1]["table_o"],
            item[1]["where_j"],
            item[1]["select_o"],
        ),
        reverse=True,
    )
    return scored[:limit]


def summarize_diff(base: SqlRecord, cand: SqlRecord, metrics: Dict[str, float], label: str) -> Dict[str, str]:
    left = base.features
    right = cand.features
    common_tables = [t for t in left.tables if t in set(right.tables)]
    left_only_tables = [t for t in left.tables if t not in set(right.tables)]
    right_only_tables = [t for t in right.tables if t not in set(left.tables)]
    common_conds = [c for c in left.where_conditions if c in set(right.where_conditions)]
    left_only_conds = [c for c in left.where_conditions if c not in set(right.where_conditions)]
    right_only_conds = [c for c in right.where_conditions if c not in set(left.where_conditions)]
    common_groups = [g for g in left.group_by if g in set(right.group_by)]
    decision, confidence = describe_match(metrics["score"], metrics["table_o"])

    join_line_left = find_line_numbers(left.lines, ["JOIN"])
    join_line_right = find_line_numbers(right.lines, ["JOIN"])
    where_line_left = find_line_numbers(left.lines, ["WHERE", "AND", "EXISTS", "NOT EXISTS"])
    where_line_right = find_line_numbers(right.lines, ["WHERE", "AND", "EXISTS", "NOT EXISTS"])
    group_line_left = find_line_numbers(left.lines, ["GROUP BY", "HAVING"])
    group_line_right = find_line_numbers(right.lines, ["GROUP BY", "HAVING"])
    sub_line_left = find_line_numbers(left.lines, ["WITH", "UNION", "SELECT"])
    sub_line_right = find_line_numbers(right.lines, ["WITH", "UNION", "SELECT"])

    reasons = [
        f"核心表重合度 {metrics['table_o']:.0%}，共同核心表：{', '.join(common_tables[:8]) if common_tables else '无'}。",
        f"WHERE 条件重合度 {metrics['where_j']:.0%}，共同过滤主干 {len(common_conds)} 条。",
        f"JOIN 数量 {left.join_count}→{right.join_count}，子查询/嵌套 SELECT 约 {left.subquery_count}→{right.subquery_count}。",
    ]
    if common_groups:
        reasons.append(f"GROUP BY 维度仍有交集：{', '.join(common_groups[:5])}。")
    if left_only_tables:
        reasons.append(f"0709 独有对象：{', '.join(left_only_tables[:6])}。")
    if right_only_tables:
        reasons.append(f"{label} 独有对象：{', '.join(right_only_tables[:6])}。")
    if left_only_conds:
        reasons.append(f"0709 额外过滤：{'; '.join(left_only_conds[:3])}。")
    if right_only_conds:
        reasons.append(f"{label} 额外过滤：{'; '.join(right_only_conds[:3])}。")

    explanation = " ".join(reasons)
    if decision == "同一业务SQL":
        explanation += " 认定为同一业务 SQL 的优化/演进版本，主要差异在写法、条件补充、历史表切换或字段裁剪。"
    elif decision == "可能同一业务SQL":
        explanation += " 结构主干接近，但存在部分查询目标或过滤条件变化，建议人工复核。"
    else:
        explanation += " 共同核心表或过滤主干不足，无法可靠认定为同一业务 SQL。"

    return {
        "判断": decision,
        "置信度": confidence,
        "语义得分": round(metrics["score"], 4),
        "共同核心表": ", ".join(common_tables[:12]),
        "0709独有表": ", ".join(left_only_tables[:12]),
        f"{label}独有表": ", ".join(right_only_tables[:12]),
        "共同过滤条件数": str(len(common_conds)),
        "0709额外过滤条件": "\n".join(left_only_conds[:8]),
        f"{label}额外过滤条件": "\n".join(right_only_conds[:8]),
        "JOIN变化": f"{left.join_count} -> {right.join_count}",
        "GROUP BY变化": ("有 -> 有" if left.has_group_by and right.has_group_by else f"{left.has_group_by} -> {right.has_group_by}"),
        "子查询变化": f"{left.subquery_count} -> {right.subquery_count}",
        "0709 JOIN/WHERE/GROUP 行号": f"JOIN:{format_line_refs(join_line_left)} | WHERE:{format_line_refs(where_line_left)} | GROUP:{format_line_refs(group_line_left)} | 结构:{format_line_refs(sub_line_left)}",
        f"{label} JOIN/WHERE/GROUP 行号": f"JOIN:{format_line_refs(join_line_right)} | WHERE:{format_line_refs(where_line_right)} | GROUP:{format_line_refs(group_line_right)} | 结构:{format_line_refs(sub_line_right)}",
        "详细说明": explanation,
    }


def build_results() -> Dict[str, object]:
    records_0709 = load_0709_records()
    records_0710 = load_0710_records()
    records_0713 = load_0713_records()

    detail_rows: List[Dict[str, object]] = []
    candidate_rows: List[Dict[str, object]] = []
    formatted_rows: Dict[str, List[Dict[str, object]]] = {"0709": [], "0710": [], "0713": []}

    for rec in records_0709:
        formatted_rows["0709"].append(
            {
                "序号": rec.seq,
                "原表行号": rec.source_row,
                "任务编号": rec.meta["任务编号"],
                "代码位置": rec.meta["代码位置"],
                "格式化SQL": "\n".join(f"{idx:03d} | {line}" for idx, line in enumerate(rec.features.lines, start=1)),
            }
        )
    for bucket, rows in [("0710", records_0710), ("0713", records_0713)]:
        for rec in rows:
            formatted_rows[bucket].append(
                {
                    "序号": rec.seq,
                    "原表行号": rec.source_row,
                    "服务": rec.meta["服务"],
                    "指纹": rec.meta["指纹"],
                    "格式化SQL": "\n".join(f"{idx:03d} | {line}" for idx, line in enumerate(rec.features.lines, start=1)),
                }
            )

    for base in records_0709:
        top_0710 = choose_candidates(base, records_0710, limit=3)
        top_0713 = choose_candidates(base, records_0713, limit=3)
        best_0710, metrics_0710 = top_0710[0]
        best_0713, metrics_0713 = top_0713[0]
        summary_0710 = summarize_diff(base, best_0710, metrics_0710, "7月10")
        summary_0713 = summarize_diff(base, best_0713, metrics_0713, "7月13")

        detail_rows.append(
            {
                "0709序号": base.seq,
                "0709原表行号": base.source_row,
                "任务编号": base.meta["任务编号"],
                "应用场景": base.meta["应用场景"],
                "代码位置": base.meta["代码位置"],
                "修复情况": base.meta["修复情况"],
                "开发人员": base.meta["开发人员"],
                "0709 SQL": base.sql,
                "7月10日判断": summary_0710["判断"],
                "7月10日置信度": summary_0710["置信度"],
                "7月10日语义得分": summary_0710["语义得分"],
                "7月10日匹配序号": best_0710.seq,
                "7月10日原表行号": best_0710.source_row,
                "7月10日指纹": best_0710.meta["指纹"],
                "7月10日服务": best_0710.meta["服务"],
                "7月10日候选SQL": best_0710.sql,
                "7月10日行号标注": summary_0710["7月10 JOIN/WHERE/GROUP 行号"],
                "7月10日详细分析": summary_0710["详细说明"],
                "7月13日判断": summary_0713["判断"],
                "7月13日置信度": summary_0713["置信度"],
                "7月13日语义得分": summary_0713["语义得分"],
                "7月13日匹配序号": best_0713.seq,
                "7月13日原表行号": best_0713.source_row,
                "7月13日指纹": best_0713.meta["指纹"],
                "7月13日服务": best_0713.meta["服务"],
                "7月13日候选SQL": best_0713.sql,
                "7月13日行号标注": summary_0713["7月13 JOIN/WHERE/GROUP 行号"],
                "7月13日详细分析": summary_0713["详细说明"],
                "0709行号标注": summary_0710["0709 JOIN/WHERE/GROUP 行号"],
                "综合结论": f"7月10：{summary_0710['判断']}；7月13：{summary_0713['判断']}",
            }
        )

        for label, scored in [("7月10", top_0710), ("7月13", top_0713)]:
            for rank, (cand, metrics) in enumerate(scored, start=1):
                candidate_rows.append(
                    {
                        "0709序号": base.seq,
                        "任务编号": base.meta["任务编号"],
                        "对比日期": label,
                        "候选排名": rank,
                        "候选序号": cand.seq,
                        "候选原表行号": cand.source_row,
                        "候选指纹": cand.meta["指纹"],
                        "候选服务": cand.meta["服务"],
                        "语义得分": round(metrics["score"], 4),
                        "核心表重合度": round(metrics["table_o"], 4),
                        "WHERE重合度": round(metrics["where_j"], 4),
                        "SELECT重合度": round(metrics["select_o"], 4),
                        "JOIN数(0709->候选)": f"{base.features.join_count}->{cand.features.join_count}",
                        "子查询数(0709->候选)": f"{base.features.subquery_count}->{cand.features.subquery_count}",
                        "0709核心表": ", ".join(base.features.tables[:12]),
                        "候选核心表": ", ".join(cand.features.tables[:12]),
                        "0709 SQL": base.sql,
                        "候选 SQL": cand.sql,
                    }
                )

    summary = Counter(row["7月10日判断"] for row in detail_rows)
    summary_0713_counter = Counter(row["7月13日判断"] for row in detail_rows)
    summary_rows = [
        {"统计项": "0709待核对SQL", "数量": len(detail_rows)},
        {"统计项": "7月10日认定同一业务SQL", "数量": summary.get("同一业务SQL", 0)},
        {"统计项": "7月10日可能同一业务SQL", "数量": summary.get("可能同一业务SQL", 0)},
        {"统计项": "7月10日非同一业务SQL", "数量": summary.get("非同一业务SQL", 0)},
        {"统计项": "7月13日认定同一业务SQL", "数量": summary_0713_counter.get("同一业务SQL", 0)},
        {"统计项": "7月13日可能同一业务SQL", "数量": summary_0713_counter.get("可能同一业务SQL", 0)},
        {"统计项": "7月13日非同一业务SQL", "数量": summary_0713_counter.get("非同一业务SQL", 0)},
        {"统计项": "说明", "数量": "行号基于本次导出的格式化SQL，不是原单元格内字符串位置。"},
    ]

    return {
        "summary": summary_rows,
        "details": detail_rows,
        "candidates": candidate_rows,
        "formatted_0709": formatted_rows["0709"],
        "formatted_0710": formatted_rows["0710"],
        "formatted_0713": formatted_rows["0713"],
    }


def main() -> None:
    output = build_results()
    out_dir = ROOT / "outputs"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "semantic_sql_compare_results.json"
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out_path)


if __name__ == "__main__":
    main()
