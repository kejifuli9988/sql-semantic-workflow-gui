import json
import os
import subprocess
import argparse
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import openpyxl
import pandas as pd

from openai_sql_semantic_matcher import LlmJudgeResult, judge_pair_with_openai
from semantic_sql_compare import (
    ROOT,
    SqlRecord,
    canonicalize_sql,
    choose_candidates,
    summarize_diff,
)


BASE_DIR = ROOT / "0804"
NODE_BIN = Path("/Users/guo/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node")
WORKBOOK_BUILDER = ROOT / "scripts" / "generic_workbook_builder.mjs"


@dataclass
class CommonSqlRecord:
    seq: int
    normalized_sql: str
    sql: str
    source_723_row: int
    source_0803_row: int
    task_id: str
    scene: str
    code_location: str
    fix_status: str
    rectify_action: str
    record: SqlRecord


def read_html_xls(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="ignore")
    df = pd.read_html(StringIO(text))[0]
    first_row = [str(x).strip() for x in df.iloc[0].tolist()]
    if "SQL语句" in first_row:
        df.columns = first_row
        df = df.iloc[1:].reset_index(drop=True)
    return df


def load_723_df() -> pd.DataFrame:
    wb = openpyxl.load_workbook(BASE_DIR / "7.23修复慢sql.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


def load_0803_df() -> pd.DataFrame:
    wb = openpyxl.load_workbook(BASE_DIR / "技术治理任务明细列表_20260803_170001.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


def load_target_records(path: Path) -> List[SqlRecord]:
    df = read_html_xls(path)
    rows: List[SqlRecord] = []
    for idx, row in df.reset_index(drop=True).iterrows():
        sql = str(row.get("SQL语句") or "")
        if not sql or sql == "SQL语句":
            continue
        seq_raw = row.get("序号")
        try:
            seq = int(str(seq_raw).strip())
        except Exception:
            seq = len(rows) + 1
        rows.append(
            SqlRecord(
                seq=seq,
                source_row=idx + 2,
                sql=sql,
                meta={
                    "来源": str(row.get("来源") or ""),
                    "指纹": str(row.get("指纹") or ""),
                    "服务": str(row.get("服务") or ""),
                    "平均响应时间(ms)": str(row.get("平均响应时间(ms)") or ""),
                    "执行次数": str(row.get("执行次数") or ""),
                },
            )
        )
    return rows


def first_non_empty(*values: object) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return ""


def build_common_records() -> List[CommonSqlRecord]:
    df_723 = load_723_df()
    df_0803 = load_0803_df()

    map_723: Dict[str, Dict[str, object]] = {}
    for idx, row in df_723.reset_index(drop=True).iterrows():
        sql = str(row.get("任务详情") or "")
        if not sql:
            continue
        key = canonicalize_sql(sql).lower()
        if not key or key in map_723:
            continue
        map_723[key] = {
            "row": idx + 2,
            "任务编号": first_non_empty(row.get("任务编号")),
            "应用场景": first_non_empty(row.get("应用场景")),
            "代码位置": first_non_empty(row.get("代码位置")),
            "修复情况": first_non_empty(row.get("修复情况")),
            "SQL": sql,
        }

    map_0803: Dict[str, Dict[str, object]] = {}
    for idx, row in df_0803.reset_index(drop=True).iterrows():
        sql = str(row.get("任务详情") or "")
        if not sql:
            continue
        key = canonicalize_sql(sql).lower()
        if not key or key in map_0803:
            continue
        map_0803[key] = {
            "row": idx + 2,
            "任务编号": first_non_empty(row.get("任务编号")),
            "整改措施": first_non_empty(row.get("整改措施")),
            "备注说明": first_non_empty(row.get("备注说明")),
            "SQL": sql,
        }

    common_keys = [key for key in map_723 if key in map_0803]
    records: List[CommonSqlRecord] = []
    for seq, key in enumerate(common_keys, start=1):
        left = map_723[key]
        right = map_0803[key]
        sql = left["SQL"]
        task_id = first_non_empty(left["任务编号"], right["任务编号"])
        scene = first_non_empty(left["应用场景"])
        code_location = first_non_empty(left["代码位置"])
        fix_status = first_non_empty(left["修复情况"])
        rectify_action = first_non_empty(right["整改措施"], right["备注说明"])
        record = SqlRecord(
            seq=seq,
            source_row=int(left["row"]),
            sql=sql,
            meta={
                "任务编号": task_id,
                "应用场景": scene,
                "代码位置": code_location,
                "修复情况": fix_status,
                "整改措施": rectify_action,
            },
        )
        records.append(
            CommonSqlRecord(
                seq=seq,
                normalized_sql=key,
                sql=sql,
                source_723_row=int(left["row"]),
                source_0803_row=int(right["row"]),
                task_id=task_id,
                scene=scene,
                code_location=code_location,
                fix_status=fix_status,
                rectify_action=rectify_action,
                record=record,
            )
        )
    return records


def build_common_workbook_payload(records: Sequence[CommonSqlRecord]) -> Dict[str, object]:
    rows = []
    for rec in records:
        rows.append(
            {
                "共有序号": rec.seq,
                "任务编号": rec.task_id,
                "7.23原表行号": rec.source_723_row,
                "0803原表行号": rec.source_0803_row,
                "应用场景": rec.scene,
                "代码位置": rec.code_location,
                "7.23修复情况": rec.fix_status,
                "0803整改措施": rec.rectify_action,
                "共有SQL": rec.sql,
            }
        )
    return {
        "sheets": [
            {
                "name": "共有SQL明细",
                "title": "7.23修复慢sql 与 2026-08-03 技术治理任务明细共有 SQL",
                "columns": ["共有序号", "任务编号", "7.23原表行号", "0803原表行号", "应用场景", "代码位置", "7.23修复情况", "0803整改措施", "共有SQL"],
                "widths": [8, 24, 10, 10, 26, 34, 16, 28, 72],
                "rows": rows,
            }
        ]
    }


def map_target_summary_label(label: str) -> str:
    return label.replace("日", "")


def build_final_payload(
    common_records: Sequence[CommonSqlRecord],
    target_730: Sequence[SqlRecord],
    target_731: Sequence[SqlRecord],
    use_openai: bool = False,
) -> Tuple[Dict[str, object], List[Dict[str, object]]]:
    detail_rows: List[Dict[str, object]] = []
    candidate_rows: List[Dict[str, object]] = []

    for common in common_records:
        ranked_730 = choose_candidates(common.record, target_730, limit=3)
        ranked_731 = choose_candidates(common.record, target_731, limit=3)
        best_730, metrics_730, judge_730 = select_best_match(common.record, ranked_730, "共有SQL", "7月30日", use_openai)
        best_731, metrics_731, judge_731 = select_best_match(common.record, ranked_731, "共有SQL", "7月31日", use_openai)
        summary_730 = build_summary_from_judge(common.record, best_730, metrics_730, judge_730, "7月30")
        summary_731 = build_summary_from_judge(common.record, best_731, metrics_731, judge_731, "7月31")

        detail_rows.append(
            {
                "共有序号": common.seq,
                "7.23原表行号": common.source_723_row,
                "0803原表行号": common.source_0803_row,
                "任务编号": common.task_id,
                "应用场景": common.scene,
                "代码位置": common.code_location,
                "7.23修复情况": common.fix_status,
                "0803整改措施": common.rectify_action,
                "共有SQL": common.sql,
                "7月30日判断": summary_730["判断"],
                "7月30日置信度": summary_730["置信度"],
                "7月30日匹配序号": best_730.seq,
                "7月30日原表行号": best_730.source_row,
                "7月30日指纹": best_730.meta["指纹"],
                "7月30日服务": best_730.meta["服务"],
                "7月30日候选SQL": best_730.sql,
                "7月30日详细分析": (
                    f"共有SQL行号：{summary_730['0709 JOIN/WHERE/GROUP 行号']}\n"
                    f"7月30行号：{summary_730['7月30 JOIN/WHERE/GROUP 行号']}\n"
                    f"{summary_730['详细说明']}"
                ),
                "7月31日判断": summary_731["判断"],
                "7月31日置信度": summary_731["置信度"],
                "7月31日匹配序号": best_731.seq,
                "7月31日原表行号": best_731.source_row,
                "7月31日指纹": best_731.meta["指纹"],
                "7月31日服务": best_731.meta["服务"],
                "7月31日候选SQL": best_731.sql,
                "7月31日详细分析": (
                    f"共有SQL行号：{summary_731['0709 JOIN/WHERE/GROUP 行号']}\n"
                    f"7月31行号：{summary_731['7月31 JOIN/WHERE/GROUP 行号']}\n"
                    f"{summary_731['详细说明']}"
                ),
                "综合结论": f"7月30日：{summary_730['判断']}；7月31日：{summary_731['判断']}",
            }
        )

        for date_label, ranked in [("7月30日", ranked_730), ("7月31日", ranked_731)]:
            for rank, (cand, metrics) in enumerate(ranked, start=1):
                candidate_rows.append(
                    {
                        "共有序号": common.seq,
                        "任务编号": common.task_id,
                        "对比日期": date_label,
                        "候选排名": rank,
                        "候选序号": cand.seq,
                        "候选原表行号": cand.source_row,
                        "候选指纹": cand.meta["指纹"],
                        "候选服务": cand.meta["服务"],
                        "语义得分": round(metrics["score"], 4),
                        "核心表重合度": round(metrics["table_o"], 4),
                        "WHERE重合度": round(metrics["where_j"], 4),
                        "SELECT重合度": round(metrics["select_o"], 4),
                        "JOIN数(共有SQL->候选)": f"{common.record.features.join_count}->{cand.features.join_count}",
                        "子查询数(共有SQL->候选)": f"{common.record.features.subquery_count}->{cand.features.subquery_count}",
                        "共有SQL": common.sql,
                        "候选SQL": cand.sql,
                    }
                )

    same_730 = sum(1 for row in detail_rows if row["7月30日判断"] == "同一业务SQL")
    possible_730 = sum(1 for row in detail_rows if row["7月30日判断"] == "可能同一业务SQL")
    no_730 = sum(1 for row in detail_rows if row["7月30日判断"] in {"非同一业务SQL", "未匹配"})
    same_731 = sum(1 for row in detail_rows if row["7月31日判断"] == "同一业务SQL")
    possible_731 = sum(1 for row in detail_rows if row["7月31日判断"] == "可能同一业务SQL")
    no_731 = sum(1 for row in detail_rows if row["7月31日判断"] in {"非同一业务SQL", "未匹配"})

    payload = {
        "sheets": [
            {
                "name": "汇总",
                "title": "0804 共有SQL 与 7月30 / 7月31 慢SQL AI语义匹配汇总",
                "summaryColumns": ["统计项", "数量"],
                "summaryRows": [
                    {"统计项": "共有SQL数量", "数量": len(common_records)},
                    {"统计项": "7月30日认定同一业务SQL", "数量": same_730},
                    {"统计项": "7月30日可能同一业务SQL", "数量": possible_730},
                    {"统计项": "7月30日未匹配/非同一业务SQL", "数量": no_730},
                    {"统计项": "7月31日认定同一业务SQL", "数量": same_731},
                    {"统计项": "7月31日可能同一业务SQL", "数量": possible_731},
                    {"统计项": "7月31日未匹配/非同一业务SQL", "数量": no_731},
                    {"统计项": "说明", "数量": "行号基于本次导出的格式化SQL，不是原单元格内字符位置。"},
                ],
                "columns": ["统计项", "数量"],
                "widths": [34, 22],
                "rows": [],
            },
            {
                "name": "逐条语义匹配",
                "title": "共有SQL 逐条语义匹配",
                "columns": [
                    "共有序号",
                    "7.23原表行号",
                    "0803原表行号",
                    "任务编号",
                    "应用场景",
                    "代码位置",
                    "7.23修复情况",
                    "0803整改措施",
                    "共有SQL",
                    "7月30日判断",
                    "7月30日置信度",
                    "7月30日匹配序号",
                    "7月30日原表行号",
                    "7月30日指纹",
                    "7月30日服务",
                    "7月30日候选SQL",
                    "7月30日详细分析",
                    "7月31日判断",
                    "7月31日置信度",
                    "7月31日匹配序号",
                    "7月31日原表行号",
                    "7月31日指纹",
                    "7月31日服务",
                    "7月31日候选SQL",
                    "7月31日详细分析",
                    "综合结论",
                ],
                "widths": [8, 10, 10, 24, 24, 32, 16, 24, 42, 12, 10, 10, 10, 18, 14, 42, 56, 12, 10, 10, 10, 18, 14, 42, 56, 20],
                "rows": detail_rows,
                "highlights": [
                    {
                        "colIndex": 10,
                        "rules": [
                            {"text": "同一业务SQL", "fill": "#E2F0D9", "fontColor": "#215E21"},
                            {"text": "可能同一业务SQL", "fill": "#FFF2CC", "fontColor": "#7F6000"},
                            {"text": "非同一业务SQL", "fill": "#FCE4D6", "fontColor": "#9E480E"},
                        ],
                    },
                    {
                        "colIndex": 18,
                        "rules": [
                            {"text": "同一业务SQL", "fill": "#E2F0D9", "fontColor": "#215E21"},
                            {"text": "可能同一业务SQL", "fill": "#FFF2CC", "fontColor": "#7F6000"},
                            {"text": "非同一业务SQL", "fill": "#FCE4D6", "fontColor": "#9E480E"},
                        ],
                    },
                ],
            },
        ]
    }
    return payload, candidate_rows


def run_builder(input_json: Path, output_xlsx: Path) -> None:
    cmd = [str(NODE_BIN if NODE_BIN.exists() else "node"), str(WORKBOOK_BUILDER), str(input_json), str(output_xlsx)]
    subprocess.run(cmd, check=True, cwd=str(ROOT))


def judge_rank(judge: LlmJudgeResult) -> Tuple[int, int, float]:
    judgement_order = {
        "同一业务SQL": 3,
        "可能同一业务SQL": 2,
        "非同一业务SQL": 1,
    }
    confidence_order = {
        "极高": 4,
        "高": 3,
        "中": 2,
        "低": 1,
    }
    return (
        judgement_order.get(judge.judgement, 0),
        confidence_order.get(judge.confidence, 0),
        float(judge.semantic_score),
    )


def select_best_match(
    base: SqlRecord,
    ranked: Sequence[Tuple[SqlRecord, Dict[str, float]]],
    base_label: str,
    candidate_label: str,
    use_openai: bool,
) -> Tuple[SqlRecord, Dict[str, float], LlmJudgeResult | None]:
    if not use_openai:
        best, metrics = ranked[0]
        return best, metrics, None

    judged: List[Tuple[SqlRecord, Dict[str, float], LlmJudgeResult]] = []
    for cand, metrics in ranked:
        judge = judge_pair_with_openai(base, cand, base_label, candidate_label)
        judged.append((cand, metrics, judge))
    judged.sort(key=lambda item: judge_rank(item[2]), reverse=True)
    return judged[0]


def build_summary_from_judge(
    base: SqlRecord,
    cand: SqlRecord,
    metrics: Dict[str, float],
    judge: LlmJudgeResult | None,
    label_without_day: str,
) -> Dict[str, str]:
    if judge is None:
        return summarize_diff(base, cand, metrics, label_without_day)
    return {
        "判断": judge.judgement,
        "置信度": judge.confidence,
        "语义得分": round(float(judge.semantic_score), 4),
        "共同核心表": ", ".join(judge.common_tables),
        "0709独有表": "",
        f"{label_without_day}独有表": "",
        "共同过滤条件数": "",
        "0709额外过滤条件": "",
        f"{label_without_day}额外过滤条件": "\n".join(judge.key_differences),
        "JOIN变化": judge.join_change,
        "GROUP BY变化": judge.group_by_change,
        "子查询变化": judge.subquery_change,
        "0709 JOIN/WHERE/GROUP 行号": judge.base_line_refs,
        f"{label_without_day} JOIN/WHERE/GROUP 行号": judge.candidate_line_refs,
        "详细说明": judge.reasoning,
    }


def resolve_use_openai(match_mode: str) -> bool:
    has_key = bool(os.getenv("OPENAI_API_KEY", "").strip())
    if match_mode == "heuristic":
        return False
    if match_mode == "openai":
        if not has_key:
            raise RuntimeError("match-mode=openai 但未设置 OPENAI_API_KEY。")
        return True
    return has_key


def main() -> None:
    parser = argparse.ArgumentParser(description="0804 共有SQL + 7月30/7月31 语义比对工作流")
    parser.add_argument("--match-mode", choices=["auto", "openai", "heuristic"], default="auto")
    args = parser.parse_args()

    output_dir = ROOT / "outputs" / "0804_common_compare"
    output_dir.mkdir(parents=True, exist_ok=True)
    use_openai = resolve_use_openai(args.match_mode)

    common_records = build_common_records()
    target_730 = load_target_records(BASE_DIR / "慢SQL表_2026_7_30.xls")
    target_731 = load_target_records(BASE_DIR / "慢SQL表_2026_7_31.xls")

    common_payload = build_common_workbook_payload(common_records)
    common_json = output_dir / "common_sql_payload.json"
    common_json.write_text(json.dumps(common_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    common_xlsx = output_dir / "7.23与0803共有SQL明细.xlsx"
    run_builder(common_json, common_xlsx)

    final_payload, candidate_rows = build_final_payload(common_records, target_730, target_731, use_openai=use_openai)
    final_json = output_dir / "final_payload.json"
    final_json.write_text(json.dumps(final_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    candidate_json = output_dir / "candidate_details.json"
    candidate_json.write_text(json.dumps(candidate_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    final_xlsx = output_dir / "0804_共有SQL_7月30日_7月31日_AI人工语义匹配完整版.xlsx"
    run_builder(final_json, final_xlsx)

    print(common_xlsx)
    print(final_xlsx)
    print(candidate_json)


if __name__ == "__main__":
    main()
