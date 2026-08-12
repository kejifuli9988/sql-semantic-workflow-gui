import argparse
import json
import subprocess
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import openpyxl
import pandas as pd

from semantic_sql_compare import (
    ROOT,
    SqlRecord,
    choose_candidates,
    summarize_diff,
)


DEFAULT_NODE = Path("/Users/guo/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node")
TEMPLATE_GENERATOR = ROOT / "scripts" / "ai_match_template_generator.mjs"


def looks_like_html_excel(path: Path) -> bool:
    if path.suffix.lower() in {".html", ".htm"}:
        return True
    with path.open("rb") as f:
        head = f.read(1024).lstrip()
    return head.startswith(b"<html") or head.startswith(b"<!DOCTYPE html") or b"<table" in head[:1024].lower()


def load_html_table(path: Path) -> pd.DataFrame:
    html = path.read_text(encoding="utf-8", errors="ignore")
    tables = pd.read_html(html)
    if not tables:
        raise ValueError(f"未能从 HTML 表格文件读取数据: {path}")
    df = tables[0]
    if df.empty:
        raise ValueError(f"HTML 表格为空: {path}")
    first_row = [str(x).strip() for x in df.iloc[0].tolist()]
    if "SQL语句" in first_row:
        df.columns = first_row
        df = df.iloc[1:].reset_index(drop=True)
    return df


def load_excel_table(path: Path) -> pd.DataFrame:
    if looks_like_html_excel(path):
        return load_html_table(path)
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(x).strip() if x is not None else "" for x in rows[0]]
        data = rows[1:]
        return pd.DataFrame(data, columns=header)
    df = pd.read_excel(path, engine="xlrd")
    return df


def pick_column(columns: Sequence[str], candidates: Sequence[str]) -> str:
    normalized = {str(col).strip(): col for col in columns}
    for name in candidates:
        if name in normalized:
            return normalized[name]
    raise KeyError(f"缺少列，期望其中之一: {candidates}，实际列: {list(columns)}")


def load_0709_records(path: Path) -> List[SqlRecord]:
    df = load_excel_table(path)
    task_col = pick_column(df.columns, ["任务编号"])
    sql_col = pick_column(df.columns, ["SQL语句"])
    scene_col = pick_column(df.columns, ["应用场景"])
    code_col = pick_column(df.columns, ["代码位置"])
    fix_col = pick_column(df.columns, ["修复情况"])
    owner_col = pick_column(df.columns, ["开发人员"])

    rows: List[SqlRecord] = []
    for idx, row in df.reset_index(drop=True).iterrows():
        if pd.isna(row.get(task_col)) and pd.isna(row.get(sql_col)):
            continue
        rows.append(
            SqlRecord(
                seq=len(rows) + 1,
                source_row=idx + 2,
                sql=str(row.get(sql_col) or ""),
                meta={
                    "任务编号": str(row.get(task_col) or ""),
                    "应用场景": str(row.get(scene_col) or ""),
                    "代码位置": str(row.get(code_col) or ""),
                    "修复情况": str(row.get(fix_col) or ""),
                    "开发人员": str(row.get(owner_col) or ""),
                },
            )
        )
    return rows


def load_target_records(path: Path) -> List[SqlRecord]:
    df = load_excel_table(path)
    seq_col = pick_column(df.columns, ["序号"])
    sql_col = pick_column(df.columns, ["SQL语句"])
    fingerprint_col = pick_column(df.columns, ["指纹"])
    service_col = pick_column(df.columns, ["服务"])
    source_col = pick_column(df.columns, ["来源"])
    count_col = pick_column(df.columns, ["执行次数"])
    latency_col = pick_column(df.columns, ["平均响应时间(ms)"])

    rows: List[SqlRecord] = []
    for idx, row in df.reset_index(drop=True).iterrows():
        if pd.isna(row.get(sql_col)):
            continue
        seq_raw = row.get(seq_col)
        try:
            seq = int(str(seq_raw).strip())
        except Exception:
            seq = len(rows) + 1
        rows.append(
            SqlRecord(
                seq=seq,
                source_row=idx + 2,
                sql=str(row.get(sql_col) or ""),
                meta={
                    "来源": str(row.get(source_col) or ""),
                    "指纹": str(row.get(fingerprint_col) or ""),
                    "服务": str(row.get(service_col) or ""),
                    "平均响应时间(ms)": str(row.get(latency_col) or ""),
                    "执行次数": str(row.get(count_col) or ""),
                },
            )
        )
    return rows


def map_target_columns(label: str) -> Dict[str, str]:
    label = label.replace(" ", "")
    if "13" in label:
        return {
            "判断": "7月13日判断",
            "置信度": "7月13日置信度",
            "匹配序号": "7月13日匹配序号",
            "原表行号": "7月13日原表行号",
            "指纹": "7月13日指纹",
            "服务": "7月13日服务",
            "候选SQL": "7月13日候选SQL",
            "详细分析": "7月13日详细分析",
        }
    return {
        "判断": "7月10日判断",
        "置信度": "7月10日置信度",
        "匹配序号": "7月10日匹配序号",
        "原表行号": "7月10日原表行号",
        "指纹": "7月10日指纹",
        "服务": "7月10日服务",
        "候选SQL": "7月10日候选SQL",
        "详细分析": "7月10日详细分析",
    }


def build_workflow_output(
    records_0709: Sequence[SqlRecord],
    target_records: Sequence[SqlRecord],
    target_label: str,
    candidate_limit: int,
) -> Dict[str, object]:
    target_map = map_target_columns(target_label)
    detail_rows: List[Dict[str, object]] = []
    candidate_rows: List[Dict[str, object]] = []

    for base in records_0709:
        ranked = choose_candidates(base, target_records, limit=candidate_limit)
        best, metrics = ranked[0]
        summary = summarize_diff(base, best, metrics, target_label.replace("日", ""))

        row = {
            "0709序号": base.seq,
            "0709原表行号": base.source_row,
            "任务编号": base.meta["任务编号"],
            "应用场景": base.meta["应用场景"],
            "代码位置": base.meta["代码位置"],
            "原修复情况": base.meta["修复情况"],
            "0709 SQL": base.sql,
            "7月10日判断": "",
            "7月10日置信度": "",
            "7月10日匹配序号": "",
            "7月10日原表行号": "",
            "7月10日指纹": "",
            "7月10日服务": "",
            "7月10日候选SQL": "",
            "7月10日详细分析": "",
            "7月13日判断": "",
            "7月13日置信度": "",
            "7月13日匹配序号": "",
            "7月13日原表行号": "",
            "7月13日指纹": "",
            "7月13日服务": "",
            "7月13日候选SQL": "",
            "7月13日详细分析": "",
            "综合结论": f"{target_label}：{summary['判断']}",
        }
        row[target_map["判断"]] = summary["判断"]
        row[target_map["置信度"]] = summary["置信度"]
        row[target_map["匹配序号"]] = best.seq
        row[target_map["原表行号"]] = best.source_row
        row[target_map["指纹"]] = best.meta["指纹"]
        row[target_map["服务"]] = best.meta["服务"]
        row[target_map["候选SQL"]] = best.sql
        row[target_map["详细分析"]] = (
            f"0709行号：{summary['0709 JOIN/WHERE/GROUP 行号']}\n"
            f"{target_label}行号：{summary[f'{target_label.replace('日', '')} JOIN/WHERE/GROUP 行号']}\n"
            f"{summary['详细说明']}"
        )
        detail_rows.append(row)

        for rank, (cand, item_metrics) in enumerate(ranked, start=1):
            candidate_rows.append(
                {
                    "0709序号": base.seq,
                    "任务编号": base.meta["任务编号"],
                    "对比日期": target_label,
                    "候选排名": rank,
                    "候选序号": cand.seq,
                    "候选原表行号": cand.source_row,
                    "候选指纹": cand.meta["指纹"],
                    "候选服务": cand.meta["服务"],
                    "语义得分": round(item_metrics["score"], 4),
                    "核心表重合度": round(item_metrics["table_o"], 4),
                    "WHERE重合度": round(item_metrics["where_j"], 4),
                    "SELECT重合度": round(item_metrics["select_o"], 4),
                    "JOIN数(0709->候选)": f"{base.features.join_count}->{cand.features.join_count}",
                    "子查询数(0709->候选)": f"{base.features.subquery_count}->{cand.features.subquery_count}",
                    "0709 SQL": base.sql,
                    "候选 SQL": cand.sql,
                }
            )

    same_count = sum(1 for row in detail_rows if row[target_map["判断"]] == "同一业务SQL")
    possible_count = sum(1 for row in detail_rows if row[target_map["判断"]] == "可能同一业务SQL")
    no_count = sum(1 for row in detail_rows if row[target_map["判断"]] in {"非同一业务SQL", "未匹配"})
    summary_rows = [
        {"统计项": "0709待核对SQL", "数量": len(detail_rows)},
        {"统计项": f"{target_label}认定同一业务SQL", "数量": same_count},
        {"统计项": f"{target_label}可能同一业务SQL", "数量": possible_count},
        {"统计项": f"{target_label}未匹配/非同一业务SQL", "数量": no_count},
    ]
    return {
        "summary_rows": summary_rows,
        "detail_rows": detail_rows,
        "candidate_rows": candidate_rows,
    }


def export_template_excel(input_json: Path, output_xlsx: Path) -> None:
    node_bin = DEFAULT_NODE if DEFAULT_NODE.exists() else Path("node")
    cmd = [str(node_bin), str(TEMPLATE_GENERATOR), str(input_json), str(output_xlsx)]
    subprocess.run(cmd, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="0709 SQL 与目标日期 SQL 双文件 AI 语义匹配工作流")
    parser.add_argument("--base", required=True, help="0709 源文件路径")
    parser.add_argument("--target", required=True, help="目标文件路径，例如 7月10")
    parser.add_argument("--target-label", default="7月10日", help="目标日期标签，默认 7月10日")
    parser.add_argument("--output-dir", default=str(ROOT / "outputs" / "two_file_workflow"), help="输出目录")
    parser.add_argument("--candidate-limit", type=int, default=3, help="每条 0709 SQL 保留的候选数量")
    args = parser.parse_args()

    base_path = Path(args.base).resolve()
    target_path = Path(args.target).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    records_0709 = load_0709_records(base_path)
    target_records = load_target_records(target_path)
    output = build_workflow_output(records_0709, target_records, args.target_label, args.candidate_limit)

    json_path = output_dir / "ai_match_template_input.json"
    json_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    candidate_path = output_dir / "candidate_details.json"
    candidate_path.write_text(json.dumps(output["candidate_rows"], ensure_ascii=False, indent=2), encoding="utf-8")

    safe_label = args.target_label.replace(" ", "").replace("/", "_")
    xlsx_path = output_dir / f"0709_SQL_{safe_label}_AI人工语义匹配完整版.xlsx"
    export_template_excel(json_path, xlsx_path)

    print(json_path)
    print(candidate_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
