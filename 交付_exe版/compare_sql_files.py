#!/usr/bin/env python3
import argparse
import json
import re
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

SQL_KEYWORDS = [
    "WITH",
    "SELECT",
    "FROM",
    "LEFT JOIN",
    "RIGHT JOIN",
    "INNER JOIN",
    "FULL JOIN",
    "JOIN",
    "WHERE",
    "GROUP BY",
    "HAVING",
    "UNION ALL",
    "UNION",
    "ORDER BY",
    "EXISTS",
    "NOT EXISTS",
    "ON",
    "AND",
    "OR",
]

DEFAULT_SQL_COLUMNS = ["SQL语句", "SQL", "sql", "任务详情", "慢SQL"]
VALID_JUDGEMENTS = {"同一业务SQL", "可能同一业务SQL", "非同一业务SQL"}
VALID_CONFIDENCE = {"极高", "高", "中", "低"}


def strip_comments(sql: str) -> str:
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.S)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_sql(sql: str) -> str:
    return normalize_text(strip_comments(sql or "").replace("\u3000", " "))


def split_top_level(expr: str, delimiter: str) -> List[str]:
    parts: List[str] = []
    current: List[str] = []
    depth = 0
    quote = ""
    i = 0
    upper = expr.upper()
    delimiter_upper = delimiter.upper()
    while i < len(expr):
        ch = expr[i]
        if quote:
            current.append(ch)
            if ch == quote:
                quote = ""
            i += 1
            continue
        if ch in ("'", '"'):
            quote = ch
            current.append(ch)
            i += 1
            continue
        if ch == "(":
            depth += 1
            current.append(ch)
            i += 1
            continue
        if ch == ")":
            depth = max(0, depth - 1)
            current.append(ch)
            i += 1
            continue
        if depth == 0 and upper.startswith(delimiter_upper, i):
            before_ok = i == 0 or not upper[i - 1].isalnum()
            after_idx = i + len(delimiter_upper)
            after_ok = after_idx >= len(expr) or not upper[after_idx].isalnum()
            if before_ok and after_ok:
                piece = "".join(current).strip()
                if piece:
                    parts.append(piece)
                current = []
                i += len(delimiter_upper)
                continue
        current.append(ch)
        i += 1
    piece = "".join(current).strip()
    if piece:
        parts.append(piece)
    return parts


def extract_clause(sql: str, start: str, ends: Sequence[str]) -> str:
    upper = sql.upper()
    start_idx = upper.find(start)
    if start_idx < 0:
        return ""
    start_idx += len(start)
    end_idx = len(sql)
    for end in ends:
        idx = upper.find(end, start_idx)
        if idx >= 0:
            end_idx = min(end_idx, idx)
    return sql[start_idx:end_idx].strip()


def dedupe(items: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for item in items:
        key = item.lower()
        if not item or key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def extract_tables(sql: str) -> List[str]:
    matches = re.findall(r"\b(?:FROM|JOIN|UPDATE|INTO)\s+([A-Z0-9_.$#]+)", sql.upper())
    return dedupe([item.lower().split(".")[-1] for item in matches])


def extract_select_targets(sql: str) -> List[str]:
    clause = extract_clause(sql, "SELECT", [" FROM"])
    if not clause:
        return []
    clause = re.sub(r"^\s*DISTINCT\s+", "", clause, flags=re.I)
    items = [normalize_text(x).lower() for x in split_top_level(clause, ",")]
    items = [re.sub(r"\bas\s+[a-z0-9_.$#]+\b", "", item) for item in items]
    items = [re.sub(r"\b[a-z0-9_.$#]+\.", "", item) for item in items]
    return dedupe(items)


def extract_group_by(sql: str) -> List[str]:
    clause = extract_clause(sql, "GROUP BY", [" HAVING", " ORDER BY", " UNION", " UNION ALL"])
    if not clause:
        return []
    return dedupe([normalize_text(x).lower() for x in split_top_level(clause, ",")])


def extract_where_conditions(sql: str) -> List[str]:
    clause = extract_clause(sql, "WHERE", [" GROUP BY", " HAVING", " ORDER BY", " UNION", " UNION ALL"])
    if not clause:
        return []
    items = []
    for item in split_top_level(clause, "AND"):
        normalized = normalize_text(item).lower()
        normalized = re.sub(r"\b\d+\b", "?", normalized)
        normalized = re.sub(r"'[^']*'", "'?'", normalized)
        items.append(normalized)
    return dedupe(items)


def count_keyword(sql: str, pattern: str) -> int:
    return len(re.findall(pattern, sql, flags=re.I))


def format_sql_lines(sql: str) -> List[str]:
    text = strip_comments(sql or "").replace("\r", "\n").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    for keyword in sorted(SQL_KEYWORDS, key=len, reverse=True):
        pattern = re.compile(rf"\b{re.escape(keyword)}\b", flags=re.I)
        text = pattern.sub(lambda m: "\n" + m.group(0).upper(), text)
    return [normalize_text(line) for line in text.split("\n") if normalize_text(line)]


def numbered_sql(sql: str, max_lines: int = 220) -> str:
    return "\n".join(f"{idx:03d} | {line}" for idx, line in enumerate(format_sql_lines(sql)[:max_lines], start=1))


def jaccard(a: Sequence[str], b: Sequence[str]) -> float:
    sa = {x for x in a if x}
    sb = {x for x in b if x}
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def overlap(a: Sequence[str], b: Sequence[str]) -> float:
    sa = {x for x in a if x}
    sb = {x for x in b if x}
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / max(1, min(len(sa), len(sb)))


@dataclass
class SqlFeatures:
    tables: List[str]
    select_targets: List[str]
    where_conditions: List[str]
    group_by: List[str]
    join_count: int
    subquery_count: int
    union_count: int
    has_group_by: bool
    has_with: bool

    @classmethod
    def from_sql(cls, sql: str) -> "SqlFeatures":
        normalized = normalize_sql(sql)
        return cls(
            tables=extract_tables(normalized),
            select_targets=extract_select_targets(normalized),
            where_conditions=extract_where_conditions(normalized),
            group_by=extract_group_by(normalized),
            join_count=count_keyword(normalized, r"\bJOIN\b"),
            subquery_count=max(0, count_keyword(normalized, r"\bSELECT\b") - 1),
            union_count=count_keyword(normalized, r"\bUNION\b"),
            has_group_by="GROUP BY" in normalized.upper(),
            has_with=normalized.upper().startswith("WITH ") or " WITH " in normalized.upper(),
        )


@dataclass
class SqlRecord:
    source_row: int
    sql: str
    meta: Dict[str, str]
    features: SqlFeatures = field(init=False)

    def __post_init__(self) -> None:
        self.features = SqlFeatures.from_sql(self.sql)


def read_html_xls(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="ignore")
    tables = pd.read_html(StringIO(text))
    if not tables:
        raise ValueError(f"未能从文件读取 HTML 表格: {path}")
    df = tables[0]
    if not df.empty:
        first_row = [str(x).strip() for x in df.iloc[0].tolist()]
        if any(name in first_row for name in DEFAULT_SQL_COLUMNS):
            df.columns = first_row
            df = df.iloc[1:].reset_index(drop=True)
    return df


def load_table(path: Path) -> pd.DataFrame:
    head = path.read_bytes()[:1024].lstrip().lower()
    if path.suffix.lower() in {".html", ".htm"} or head.startswith(b"<html") or b"<table" in head:
        return read_html_xls(path)
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(row_iter)
        except StopIteration:
            wb.close()
            return pd.DataFrame()

        header = [str(x).strip() if x is not None else "" for x in first_row]
        data_rows = [tuple(row) for row in row_iter]
        wb.close()
        return pd.DataFrame(data_rows, columns=header)
    return pd.read_excel(path, engine="xlrd")


def pick_sql_column(columns: Sequence[str], preferred: str | None) -> str:
    names = [str(col).strip() for col in columns]
    lookup = {name: name for name in names}
    if preferred:
        if preferred in lookup:
            return preferred
        raise KeyError(f"指定 SQL 列不存在: {preferred}，实际列: {names}")
    for name in DEFAULT_SQL_COLUMNS:
        if name in lookup:
            return name
    raise KeyError(f"未自动识别 SQL 列，实际列: {names}")


def load_records(
    path: Path,
    sql_column: str | None,
    extra_meta_columns: Sequence[str] | None = None,
) -> Tuple[List[SqlRecord], str]:
    df = load_table(path)
    sql_col = pick_sql_column(df.columns, sql_column)
    extra_meta_columns = list(extra_meta_columns or [])
    rows: List[SqlRecord] = []
    for idx, row in df.reset_index(drop=True).iterrows():
        sql = str(row.get(sql_col) or "").strip()
        if not sql or sql.lower() == "nan":
            continue
        meta: Dict[str, str] = {}
        for column in extra_meta_columns:
            if column in df.columns and column != sql_col:
                meta[column] = str(row.get(column) or "")
        rows.append(SqlRecord(source_row=idx + 2, sql=sql, meta=meta))
    return rows, sql_col


def coarse_score(base: SqlRecord, target: SqlRecord) -> Dict[str, float]:
    table_o = overlap(base.features.tables, target.features.tables)
    where_j = jaccard(base.features.where_conditions, target.features.where_conditions)
    select_o = overlap(base.features.select_targets, target.features.select_targets)
    group_j = jaccard(base.features.group_by, target.features.group_by)
    join_penalty = abs(base.features.join_count - target.features.join_count)
    subquery_penalty = abs(base.features.subquery_count - target.features.subquery_count)
    score = (
        table_o * 0.45
        + where_j * 0.2
        + select_o * 0.15
        + group_j * 0.1
        + (1.0 if base.features.has_group_by == target.features.has_group_by else 0.0) * 0.05
        + (1.0 if base.features.has_with == target.features.has_with else 0.0) * 0.05
        - min(join_penalty, 5) * 0.02
        - min(subquery_penalty, 5) * 0.02
    )
    return {
        "score": round(max(0.0, min(score, 1.0)), 6),
        "table_overlap": round(table_o, 6),
        "where_jaccard": round(where_j, 6),
        "select_overlap": round(select_o, 6),
        "group_jaccard": round(group_j, 6),
    }


def build_target_table_index(targets: Sequence[SqlRecord]) -> Dict[str, List[SqlRecord]]:
    index: Dict[str, List[SqlRecord]] = {}
    for target in targets:
        for table in target.features.tables:
            index.setdefault(table, []).append(target)
    return index


def prefilter_targets_by_tables(
    base: SqlRecord,
    targets: Sequence[SqlRecord],
    table_index: Dict[str, List[SqlRecord]],
) -> List[SqlRecord]:
    if not base.features.tables:
        return list(targets)

    matched: Dict[int, SqlRecord] = {}
    for table in base.features.tables:
        for target in table_index.get(table, []):
            matched[id(target)] = target

    if matched:
        return list(matched.values())
    return list(targets)


def choose_candidates(
    base: SqlRecord,
    targets: Sequence[SqlRecord],
    limit: int,
    table_index: Dict[str, List[SqlRecord]] | None = None,
) -> List[Tuple[SqlRecord, Dict[str, float]]]:
    candidate_pool = (
        prefilter_targets_by_tables(base, targets, table_index)
        if table_index is not None
        else list(targets)
    )
    ranked = [(target, coarse_score(base, target)) for target in candidate_pool]
    ranked.sort(key=lambda item: item[1]["score"], reverse=True)
    return ranked[:limit]


def expected_result_schema() -> Dict[str, object]:
    return {
        "judgement": ["同一业务SQL", "可能同一业务SQL", "非同一业务SQL"],
        "confidence": ["极高", "高", "中", "低"],
        "semantic_score": "0 到 1 的数值",
        "same_business": "布尔值",
        "reasoning": "中文详细说明",
        "join_change": "JOIN 结构变化总结",
        "where_change": "WHERE / 过滤条件变化总结",
        "group_by_change": "GROUP BY 或聚合口径变化总结",
        "subquery_change": "子查询、CTE、UNION 变化总结",
        "base_line_refs": "主 SQL 主要依据行号",
        "target_line_refs": "对比 SQL 主要依据行号",
        "common_tables": "共同核心表列表",
        "key_differences": "关键差异点列表",
    }


def feature_summary(record: SqlRecord) -> Dict[str, object]:
    return {
        "tables": record.features.tables,
        "join_count": record.features.join_count,
        "where_conditions": record.features.where_conditions[:15],
        "group_by": record.features.group_by,
        "subquery_count": record.features.subquery_count,
        "union_count": record.features.union_count,
    }


def build_review_prompt(base: SqlRecord, target: SqlRecord, coarse: Dict[str, float]) -> str:
    return f"""
你是资深 SQL 语义分析专家。你的任务不是看字符串像不像，而是判断两条 SQL 是否属于同一业务 SQL。

请严格按照以下标准判断：
1. 重点看查询目标、核心表、JOIN 关系、WHERE 主干、GROUP BY 口径、子查询/CTE/UNION 的业务含义是否一致。
2. 如果只是改写、分页变化、条件下推、历史分表、字段增减、参数字面值差异，通常仍可判为同一业务SQL或可能同一业务SQL。
3. 如果查询目标、统计口径、核心过滤逻辑明显不同，就判为非同一业务SQL。
4. 必须引用我提供的带行号 SQL，指出你主要依据的具体行号。
5. 请用中文输出详细说明。

你必须输出这些字段：
- judgement
- confidence
- semantic_score
- same_business
- reasoning
- join_change
- where_change
- group_by_change
- subquery_change
- base_line_refs
- target_line_refs
- common_tables
- key_differences

候选粗召回信息：
{json.dumps(coarse, ensure_ascii=False)}

主文件 SQL 结构特征：
{json.dumps({
    "tables": base.features.tables,
    "join_count": base.features.join_count,
    "where_conditions": base.features.where_conditions[:15],
    "group_by": base.features.group_by,
    "subquery_count": base.features.subquery_count,
    "union_count": base.features.union_count,
}, ensure_ascii=False)}

对比文件 SQL 结构特征：
{json.dumps({
    "tables": target.features.tables,
    "join_count": target.features.join_count,
    "where_conditions": target.features.where_conditions[:15],
    "group_by": target.features.group_by,
    "subquery_count": target.features.subquery_count,
    "union_count": target.features.union_count,
}, ensure_ascii=False)}

主文件 SQL（带行号）：
{numbered_sql(base.sql)}

对比文件 SQL（带行号）：
{numbered_sql(target.sql)}
""".strip()


def build_review_prompt_from_task(task: Dict[str, object]) -> str:
    return f"""
你是资深 SQL 语义分析专家。你的任务不是看字符串像不像，而是判断两条 SQL 是否属于同一业务 SQL。

请严格按照以下标准判断：
1. 重点看查询目标、核心表、JOIN 关系、WHERE 主干、GROUP BY 口径、子查询/CTE/UNION 的业务含义是否一致。
2. 如果只是改写、分页变化、条件下推、历史分表、字段增减、参数字面值差异，通常仍可判为同一业务SQL或可能同一业务SQL。
3. 如果查询目标、统计口径、核心过滤逻辑明显不同，就判为非同一业务SQL。
4. 必须引用我提供的带行号 SQL，指出你主要依据的具体行号。
5. 请用中文输出详细说明。

你必须输出这些字段：
- judgement
- confidence
- semantic_score
- same_business
- reasoning
- join_change
- where_change
- group_by_change
- subquery_change
- base_line_refs
- target_line_refs
- common_tables
- key_differences

候选粗召回信息：
{json.dumps(task["coarse"], ensure_ascii=False)}

主文件 SQL 结构特征：
{json.dumps(task["base_features"], ensure_ascii=False)}

对比文件 SQL 结构特征：
{json.dumps(task["target_features"], ensure_ascii=False)}

主文件 SQL（带行号）：
{numbered_sql(str(task["base_sql"]))}

对比文件 SQL（带行号）：
{numbered_sql(str(task["target_sql"]))}
""".strip()


def build_prepare_payload(
    base_path: Path,
    target_path: Path,
    base_records: Sequence[SqlRecord],
    target_records: Sequence[SqlRecord],
    base_sql_col: str,
    target_sql_col: str,
    base_display_columns: Sequence[str] | None,
    target_display_columns: Sequence[str] | None,
    top_k: int,
) -> Dict[str, object]:
    review_tasks: List[Dict[str, object]] = []
    candidate_rows: List[Dict[str, object]] = []
    pair_seq = 0
    target_table_index = build_target_table_index(target_records)
    base_display_columns = list(base_display_columns or [])
    target_display_columns = list(target_display_columns or [])

    for base in base_records:
        ranked = choose_candidates(base, target_records, max(1, top_k), target_table_index)
        for rank, (target, coarse) in enumerate(ranked, start=1):
            pair_seq += 1
            pair_id = f"pair_{pair_seq}_base_{base.source_row}_target_{target.source_row}_rank_{rank}"
            review_tasks.append(
                {
                    "pair_id": pair_id,
                    "base_row": base.source_row,
                    "base_display_values": {column: base.meta.get(column, "") for column in base_display_columns},
                    "target_row": target.source_row,
                    "target_display_values": {column: target.meta.get(column, "") for column in target_display_columns},
                    "candidate_rank": rank,
                    "base_sql": base.sql,
                    "target_sql": target.sql,
                    "base_features": feature_summary(base),
                    "target_features": feature_summary(target),
                    "coarse": coarse,
                }
            )
            candidate_rows.append(
                {
                    "pair_id": pair_id,
                    "base_row": base.source_row,
                    "base_display_values": {column: base.meta.get(column, "") for column in base_display_columns},
                    "target_row": target.source_row,
                    "target_display_values": {column: target.meta.get(column, "") for column in target_display_columns},
                    "candidate_rank": rank,
                    "coarse_score": coarse["score"],
                    "table_overlap": coarse["table_overlap"],
                    "where_jaccard": coarse["where_jaccard"],
                    "select_overlap": coarse["select_overlap"],
                    "group_jaccard": coarse["group_jaccard"],
                    "base_sql": base.sql,
                    "target_sql": target.sql,
                }
            )

    if len({item["pair_id"] for item in review_tasks}) != len(review_tasks):
        raise ValueError("prepare 阶段生成了重复 pair_id，请检查输入数据或 pair_id 规则")

    return {
        "workflow_stage": "prepare",
        "base_file": str(base_path.resolve()),
        "target_file": str(target_path.resolve()),
        "base_sql_column": base_sql_col,
        "target_sql_column": target_sql_col,
        "base_display_columns": base_display_columns,
        "target_display_columns": target_display_columns,
        "top_k": top_k,
        "base_count": len(base_records),
        "target_count": len(target_records),
        "review_tasks": review_tasks,
        "candidates": candidate_rows,
    }


def split_prepare_payload(prepared: Dict[str, object], batch_size: int) -> List[Dict[str, object]]:
    if batch_size <= 0:
        raise ValueError("batch_size 必须大于 0")

    review_tasks = list(prepared["review_tasks"])
    candidate_map: Dict[str, Dict[str, object]] = {str(item["pair_id"]): item for item in prepared["candidates"]}
    parts: List[Dict[str, object]] = []

    for start in range(0, len(review_tasks), batch_size):
        source_tasks = review_tasks[start:start + batch_size]
        chunk_tasks: List[Dict[str, object]] = []
        for task in source_tasks:
            task_copy = dict(task)
            task_copy["prompt"] = build_review_prompt_from_task(task_copy)
            task_copy["expected_result_schema"] = expected_result_schema()
            chunk_tasks.append(task_copy)
        chunk_pair_ids = {str(item["pair_id"]) for item in chunk_tasks}
        chunk_candidates = [candidate_map[pair_id] for pair_id in chunk_pair_ids if pair_id in candidate_map]
        part_no = len(parts) + 1
        parts.append(
            {
                **prepared,
                "workflow_stage": "prepare_part",
                "part_no": part_no,
                "part_count": 0,
                "batch_size": batch_size,
                "review_tasks": chunk_tasks,
                "candidates": chunk_candidates,
            }
        )

    part_count = len(parts)
    for part in parts:
        part["part_count"] = part_count
    return parts


def load_json(path: Path) -> Dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_review_results(raw: object) -> List[Dict[str, object]]:
    if isinstance(raw, dict):
        if isinstance(raw.get("results"), list):
            items = raw["results"]
        elif isinstance(raw.get("review_results"), list):
            items = raw["review_results"]
        else:
            raise ValueError("review 结果文件缺少 results 或 review_results 数组")
    elif isinstance(raw, list):
        items = raw
    else:
        raise ValueError("review 结果文件格式错误")
    return [dict(item) for item in items]


def merge_review_result_files(paths: Sequence[Path]) -> List[Dict[str, object]]:
    merged: List[Dict[str, object]] = []
    seen_pair_ids = set()
    for path in paths:
        raw = load_json(path)
        items = normalize_review_results(raw)
        for item in items:
            pair_id = str(item.get("pair_id") or "").strip()
            if not pair_id:
                raise ValueError(f"{path} 中存在缺少 pair_id 的结果")
            if pair_id in seen_pair_ids:
                raise ValueError(f"发现重复 pair_id: {pair_id}")
            seen_pair_ids.add(pair_id)
            merged.append(dict(item))
    return merged


def validate_result_item(item: Dict[str, object]) -> Dict[str, object]:
    pair_id = str(item.get("pair_id") or "").strip()
    if not pair_id:
        raise ValueError("review 结果缺少 pair_id")

    judgement = str(item.get("judgement") or "").strip()
    confidence = str(item.get("confidence") or "").strip()
    if judgement and judgement not in VALID_JUDGEMENTS:
        raise ValueError(f"{pair_id} 的 judgement 非法: {judgement}")
    if confidence and confidence not in VALID_CONFIDENCE:
        raise ValueError(f"{pair_id} 的 confidence 非法: {confidence}")

    semantic_score = item.get("semantic_score", 0)
    try:
        semantic_score = float(semantic_score)
    except Exception as exc:
        raise ValueError(f"{pair_id} 的 semantic_score 非法") from exc

    common_tables = item.get("common_tables") or []
    key_differences = item.get("key_differences") or []
    if not isinstance(common_tables, list):
        common_tables = [str(common_tables)]
    if not isinstance(key_differences, list):
        key_differences = [str(key_differences)]

    return {
        "pair_id": pair_id,
        "judgement": judgement,
        "confidence": confidence,
        "semantic_score": semantic_score,
        "same_business": bool(item.get("same_business", False)),
        "reasoning": str(item.get("reasoning") or ""),
        "join_change": str(item.get("join_change") or ""),
        "where_change": str(item.get("where_change") or ""),
        "group_by_change": str(item.get("group_by_change") or ""),
        "subquery_change": str(item.get("subquery_change") or ""),
        "base_line_refs": str(item.get("base_line_refs") or ""),
        "target_line_refs": str(item.get("target_line_refs") or ""),
        "common_tables": [str(x) for x in common_tables],
        "key_differences": [str(x) for x in key_differences],
    }


def judgement_rank(judgement: str) -> int:
    if judgement == "同一业务SQL":
        return 3
    if judgement == "可能同一业务SQL":
        return 2
    if judgement == "非同一业务SQL":
        return 1
    return 0


def choose_best_result(items: List[Dict[str, object]]) -> Dict[str, object]:
    return sorted(
        items,
        key=lambda item: (
            float(item.get("semantic_score", 0)),
            judgement_rank(str(item.get("judgement") or "")),
            -int(item.get("candidate_rank", 999999)),
        ),
        reverse=True,
    )[0]


def build_finalize_payload(prepared: Dict[str, object], review_results: List[Dict[str, object]]) -> Dict[str, object]:
    task_map = {str(item["pair_id"]): item for item in prepared["review_tasks"]}
    merged_results: List[Dict[str, object]] = []
    for raw in review_results:
        result = validate_result_item(raw)
        pair_id = result["pair_id"]
        if pair_id not in task_map:
            raise ValueError(f"未知 pair_id: {pair_id}")
        task = task_map[pair_id]
        merged_results.append(
            {
                **task,
                **result,
            }
        )

    by_base: Dict[int, List[Dict[str, object]]] = {}
    for item in merged_results:
        by_base.setdefault(int(item["base_row"]), []).append(item)

    matches: List[Dict[str, object]] = []
    for base_row, items in sorted(by_base.items()):
        best = choose_best_result(items)
        matches.append(
            {
                "base_row": best["base_row"],
                "base_display_values": dict(best.get("base_display_values") or {}),
                "base_sql": best["base_sql"],
                "best_target_row": best["target_row"],
                "best_target_sql": best["target_sql"],
                "target_display_values": dict(best.get("target_display_values") or {}),
                "judgement": best["judgement"],
                "confidence": best["confidence"],
                "semantic_score": best["semantic_score"],
                "same_business": best["same_business"],
                "reasoning": best["reasoning"],
                "join_change": best["join_change"],
                "where_change": best["where_change"],
                "group_by_change": best["group_by_change"],
                "subquery_change": best["subquery_change"],
                "base_line_refs": best["base_line_refs"],
                "target_line_refs": best["target_line_refs"],
                "common_tables": best["common_tables"],
                "key_differences": best["key_differences"],
            }
        )

    candidate_rows: List[Dict[str, object]] = []
    prepared_candidate_map = {str(item["pair_id"]): item for item in prepared["candidates"]}
    for item in merged_results:
        coarse = prepared_candidate_map[item["pair_id"]]
        candidate_rows.append(
            {
                "base_row": item["base_row"],
                "base_display_values": dict(item.get("base_display_values") or {}),
                "target_row": item["target_row"],
                "target_display_values": dict(item.get("target_display_values") or {}),
                "candidate_rank": item["candidate_rank"],
                "coarse_score": coarse["coarse_score"],
                "table_overlap": coarse["table_overlap"],
                "where_jaccard": coarse["where_jaccard"],
                "select_overlap": coarse["select_overlap"],
                "group_jaccard": coarse["group_jaccard"],
                "llm_judgement": item["judgement"],
                "llm_semantic_score": item["semantic_score"],
            }
        )

    return {
        "workflow_stage": "finalize",
        "base_file": prepared["base_file"],
        "target_file": prepared["target_file"],
        "base_sql_column": prepared["base_sql_column"],
        "target_sql_column": prepared["target_sql_column"],
        "base_display_columns": prepared.get("base_display_columns", []),
        "target_display_columns": prepared.get("target_display_columns", []),
        "top_k": prepared["top_k"],
        "base_count": prepared["base_count"],
        "target_count": prepared["target_count"],
        "review_result_count": len(merged_results),
        "matches": matches,
        "candidates": candidate_rows,
    }


def autosize_worksheet(ws: openpyxl.worksheet.worksheet.Worksheet, max_width: int = 80) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, max_width)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, width)


def write_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, rows: List[Dict[str, object]], columns: List[str]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    ws.freeze_panes = "A2"
    autosize_worksheet(ws)


def build_summary_rows(payload: Dict[str, object]) -> List[Dict[str, object]]:
    matches = payload["matches"]
    same_count = sum(1 for row in matches if row["judgement"] == "同一业务SQL")
    possible_count = sum(1 for row in matches if row["judgement"] == "可能同一业务SQL")
    different_count = sum(1 for row in matches if row["judgement"] == "非同一业务SQL")
    return [
        {"统计项": "主文件路径", "值": payload["base_file"]},
        {"统计项": "对比文件路径", "值": payload["target_file"]},
        {"统计项": "主文件 SQL 列", "值": payload["base_sql_column"]},
        {"统计项": "对比文件 SQL 列", "值": payload["target_sql_column"]},
        {"统计项": "主文件 SQL 条数", "值": payload["base_count"]},
        {"统计项": "对比文件 SQL 条数", "值": payload["target_count"]},
        {"统计项": "候选召回 top_k", "值": payload["top_k"]},
        {"统计项": "review 结果条数", "值": payload.get("review_result_count", 0)},
        {"统计项": "同一业务SQL", "值": same_count},
        {"统计项": "可能同一业务SQL", "值": possible_count},
        {"统计项": "非同一业务SQL", "值": different_count},
    ]


def export_excel(payload: Dict[str, object], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("汇总")
    write_sheet(summary_ws, build_summary_rows(payload), ["统计项", "值"])

    matches_ws = wb.create_sheet("逐条语义比对")
    match_columns = [
        "主文件行号",
        "主文件SQL",
        "对比文件匹配行号",
        "对比文件匹配SQL",
        "判断结果",
        "置信度",
        "语义评分",
        "主SQL依据行号",
        "候选SQL依据行号",
        "JOIN变化",
        "WHERE变化",
        "GROUP BY变化",
        "子查询变化",
        "共同核心表",
        "关键差异点",
        "详细说明",
    ]
    base_extra_columns = [f"主文件_{str(column)}" for column in payload.get("base_display_columns", []) if str(column).strip()]
    target_extra_columns = [f"对比文件_{str(column)}" for column in payload.get("target_display_columns", []) if str(column).strip()]
    match_columns.extend(base_extra_columns)
    match_columns.extend(target_extra_columns)
    match_rows: List[Dict[str, object]] = []
    for row in payload["matches"]:
        row_data = {
            "主文件行号": row["base_row"],
            "主文件SQL": row["base_sql"],
            "对比文件匹配行号": row["best_target_row"],
            "对比文件匹配SQL": row["best_target_sql"],
            "判断结果": row["judgement"],
            "置信度": row["confidence"],
            "语义评分": row["semantic_score"],
            "主SQL依据行号": row["base_line_refs"],
            "候选SQL依据行号": row["target_line_refs"],
            "JOIN变化": row["join_change"],
            "WHERE变化": row["where_change"],
            "GROUP BY变化": row["group_by_change"],
            "子查询变化": row["subquery_change"],
            "共同核心表": "、".join(row["common_tables"]),
            "关键差异点": "\n".join(row["key_differences"]),
            "详细说明": row["reasoning"],
        }
        for column in payload.get("base_display_columns", []):
            row_data[f"主文件_{column}"] = (row.get("base_display_values") or {}).get(column, "")
        for column in payload.get("target_display_columns", []):
            row_data[f"对比文件_{column}"] = (row.get("target_display_values") or {}).get(column, "")
        match_rows.append(row_data)
    write_sheet(matches_ws, match_rows, match_columns)

    candidates_ws = wb.create_sheet("候选明细")
    candidate_columns = [
        "主文件行号",
        "对比文件候选行号",
        "候选排名",
        "粗召回得分",
        "核心表重合度",
        "WHERE重合度",
        "SELECT重合度",
        "GROUP重合度",
        "模型判断",
        "模型语义评分",
    ]
    candidate_columns.extend(base_extra_columns)
    candidate_columns.extend(target_extra_columns)
    candidate_rows: List[Dict[str, object]] = []
    for row in payload["candidates"]:
        row_data = {
            "主文件行号": row["base_row"],
            "对比文件候选行号": row["target_row"],
            "候选排名": row["candidate_rank"],
            "粗召回得分": row["coarse_score"],
            "核心表重合度": row["table_overlap"],
            "WHERE重合度": row["where_jaccard"],
            "SELECT重合度": row["select_overlap"],
            "GROUP重合度": row["group_jaccard"],
            "模型判断": row["llm_judgement"],
            "模型语义评分": row["llm_semantic_score"],
        }
        for column in payload.get("base_display_columns", []):
            row_data[f"主文件_{column}"] = (row.get("base_display_values") or {}).get(column, "")
        for column in payload.get("target_display_columns", []):
            row_data[f"对比文件_{column}"] = (row.get("target_display_values") or {}).get(column, "")
        candidate_rows.append(row_data)
    write_sheet(candidates_ws, candidate_rows, candidate_columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def prepare_mode(args: argparse.Namespace) -> None:
    base_path = Path(args.base)
    target_path = Path(args.target)
    base_display_columns = [item.strip() for item in (args.base_display_columns or "").split(",") if item.strip()]
    target_display_columns = [item.strip() for item in (args.target_display_columns or "").split(",") if item.strip()]
    base_records, base_sql_col = load_records(base_path, args.base_sql_column, base_display_columns)
    target_records, target_sql_col = load_records(target_path, args.target_sql_column, target_display_columns)

    if not base_records:
        raise RuntimeError("主文件未读取到有效 SQL")
    if not target_records:
        raise RuntimeError("对比文件未读取到有效 SQL")

    payload = build_prepare_payload(
        base_path,
        target_path,
        base_records,
        target_records,
        base_sql_col,
        target_sql_col,
        base_display_columns,
        target_display_columns,
        args.top_k,
    )
    output_path = Path(args.output)
    write_json(output_path, payload)
    print(output_path)


def split_prepare_mode(args: argparse.Namespace) -> None:
    prepared = load_json(Path(args.prepared))
    parts = split_prepare_payload(prepared, args.batch_size)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for part in parts:
        output_path = output_dir / f"prepare_part_{int(part['part_no'])}.json"
        write_json(output_path, part)
        print(output_path)


def finalize_mode(args: argparse.Namespace) -> None:
    prepared = load_json(Path(args.prepared))
    review_results = normalize_review_results(load_json(Path(args.review_results)))
    payload = build_finalize_payload(prepared, review_results)

    output_path = Path(args.output)
    if args.json_output:
        json_output_path = Path(args.json_output)
    elif output_path.suffix.lower() == ".xlsx":
        json_output_path = output_path.with_suffix(".json")
    else:
        json_output_path = output_path

    write_json(json_output_path, payload)

    if output_path.suffix.lower() == ".xlsx":
        export_excel(payload, output_path)
        print(output_path)
    else:
        print(json_output_path)


def merge_review_results_mode(args: argparse.Namespace) -> None:
    merged = merge_review_result_files([Path(path) for path in args.review_results_files])
    output_path = Path(args.output)
    payload = {"review_results": merged}
    write_json(output_path, payload)
    print(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="文件对文件 SQL 语义匹配的预处理/后处理脚本")
    parser.add_argument(
        "--mode",
        choices=["prepare", "split-prepare", "merge-review-results", "finalize"],
        required=True,
        help="prepare 产出候选对；split-prepare 按批拆分；merge-review-results 合并分批结果；finalize 根据 review 结果输出 Excel",
    )

    parser.add_argument("--base", help="主文件路径，仅 prepare 模式需要")
    parser.add_argument("--target", help="对比文件路径，仅 prepare 模式需要")
    parser.add_argument("--base-sql-column", help="主文件 SQL 列名，仅 prepare 模式需要")
    parser.add_argument("--target-sql-column", help="对比文件 SQL 列名，仅 prepare 模式需要")
    parser.add_argument("--base-display-columns", help="主文件需要在最终 Excel 末尾展示的列名，多个逗号分隔，仅 prepare 模式需要")
    parser.add_argument("--target-display-columns", help="对比文件需要在最终 Excel 末尾展示的列名，多个逗号分隔，仅 prepare 模式需要")
    parser.add_argument("--top-k", type=int, default=3, help="每条主 SQL 召回多少个候选，仅 prepare 模式需要")
    parser.add_argument("--prepared", help="prepare 阶段产出的 JSON，split-prepare 和 finalize 模式需要")
    parser.add_argument("--batch-size", type=int, default=20, help="split-prepare 模式每批多少条，默认 20")
    parser.add_argument("--output-dir", help="split-prepare 模式输出目录")

    parser.add_argument("--review-results", help="智能体返回的结构化结果 JSON，仅 finalize 模式需要")
    parser.add_argument("--review-results-files", nargs="+", help="merge-review-results 模式输入的多个 review_results_part_x.json")
    parser.add_argument("--output", required=True, help="prepare 模式输出候选 JSON；finalize 模式推荐传 .xlsx")
    parser.add_argument("--json-output", help="仅 finalize 模式可选：显式指定 JSON 底稿输出路径")
    args = parser.parse_args()

    if args.mode == "prepare":
        if not args.base or not args.target:
            raise RuntimeError("prepare 模式必须提供 --base 和 --target")
        prepare_mode(args)
        return

    if args.mode == "split-prepare":
        if not args.prepared or not args.output_dir:
            raise RuntimeError("split-prepare 模式必须提供 --prepared 和 --output-dir")
        split_prepare_mode(args)
        return

    if args.mode == "merge-review-results":
        if not args.review_results_files:
            raise RuntimeError("merge-review-results 模式必须提供 --review-results-files")
        merge_review_results_mode(args)
        return

    if not args.prepared or not args.review_results:
        raise RuntimeError("finalize 模式必须提供 --prepared 和 --review-results")
    finalize_mode(args)


if __name__ == "__main__":
    main()
